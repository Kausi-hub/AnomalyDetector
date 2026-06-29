import re
import json
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO
from difflib import get_close_matches

import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# ==============================
# CONFIG
# ==============================
TIME_THRESHOLD_SEC = 300

# ==============================
# HELPERS
# ==============================
def normalize(name):
    name = name.lower()
    name = re.sub(r".*::", "", name)
    name = re.sub(r"[_\s]+", "", name)
    return name

def category(name):
    n = name.lower()
    if "speed" in n: return "Speed"
    if "torque" in n: return "Torque"
    if "temp" in n: return "Temperature"
    if "volt" in n: return "Voltage"
    return "Other"

def ryg(val):
    return "🟢" if val >= 75 else "🟡" if val >= 40 else "🔴"

# ==============================
# PARSERS
# ==============================
def parse_csv(file):
    lines = file.getvalue().decode(errors="ignore").splitlines()
    start, dur, reading = None, None, False
    headers, data = [], []

    for line in lines:
        if "StartMeasDateTime" in line:
            start = datetime.strptime(line.split(",")[1].strip(), "%Y/%m/%d %H:%M:%S")
        if "TotalMeasTime" in line:
            dur = float(line.split(",")[1])
        if line.startswith("Time,"):
            headers = line.split(",")
            reading = True
            continue
        if reading and "," in line:
            data.append(line.split(","))

    if not start or not dur or not data:
        return None

    df = pd.DataFrame(data, columns=headers).apply(pd.to_numeric, errors="coerce")
    signals = {normalize(c): {"values": df[c].dropna().values} for c in df.columns}

    return {"file": file.name, "start": start, "end": start + timedelta(minutes=dur), "signals": signals}

def parse_txt(file):
    lines = file.getvalue().decode(errors="ignore").splitlines()
    start, last = None, 0
    signals = defaultdict(list)

    for line in lines:
        if line.lower().startswith("date"):
            try:
                start = datetime.strptime(line.replace("date", "").strip(), "%a %b %d %I:%M:%S.%f %p %Y")
            except:
                pass

        t = re.match(r"\s*(\d+\.\d+)", line)
        if t: last = float(t.group(1))

        m = re.search(r"::([\w]+)\s*=\s*([-\d\.]+)", line)
        if m:
            signals[normalize(m.group(1))].append(float(m.group(2)))

    if not start:
        return None

    return {"file": file.name, "start": start, "end": start + timedelta(seconds=last), "signals": signals}

# ==============================
# MATCHING HELPERS
# ==============================
def apply_manual_mappings(csv_signals, txt_signals, mappings):
    return [(c, t) for c, t in mappings.items()
            if c in csv_signals and t in txt_signals]

def find_pairs(csv, txt):
    return [(c, t) for c in csv for t in txt if c in t or t in c]

def suggest_mappings(csv_signals, txt_signals):
    suggestions = {}
    for c in csv_signals:
        match = get_close_matches(c, txt_signals, n=1, cutoff=0.6)
        if match:
            suggestions[c] = match[0]
    return suggestions

def signal_similarity(csv, txt, pairs):
    scores = []
    for c, t in pairs:
        a, b = csv[c]["values"], txt[t]
        if len(a) < 10 or len(b) < 10: continue
        m = min(len(a), len(b))
        scores.append(pd.Series(a[:m]).corr(pd.Series(b[:m])))
    return round(max(0, sum(scores) / len(scores)) * 100, 1) if scores else 0

def best_signals(csv, txt, pairs):
    scored = []
    for c, t in pairs:
        a, b = csv[c]["values"], txt[t]
        if len(a) < 10 or len(b) < 10: continue
        m = min(len(a), len(b))
        corr = pd.Series(a[:m]).corr(pd.Series(b[:m]))
        if pd.notna(corr):
            scored.append((c, t, abs(corr)))
    scored.sort(key=lambda x: x[2], reverse=True)
    return [(c, t) for c, t, _ in scored[:3]]

def confidence(diff, sig):
    return round(min(100, 0.5 * (100 - diff / 300 * 100) + 0.5 * sig), 1)

# ==============================
# ANALYSIS
# ==============================
def run(csv_runs, txt_runs):
    rows, matches = [], []

    for txt in txt_runs:
        best, best_score = None, -1

        for csv in csv_runs:
            diff = abs((txt["start"] - csv["start"]).total_seconds())

            mappings = st.session_state.get("signal_map", {})
            pairs = apply_manual_mappings(csv["signals"], txt["signals"], mappings)

            if not pairs:
                pairs = find_pairs(csv["signals"], txt["signals"])

            sig = signal_similarity(csv["signals"], txt["signals"], pairs)
            conf = confidence(diff, sig)

            if best is None or conf > best_score:
                best = (csv, diff, conf, sig, pairs)
                best_score = conf

        if best is None: continue

        csv, diff, conf, sig, pairs = best

        matches.append({"csv": csv, "txt": txt, "pairs": pairs})

        rows.append({
            "TXT File": txt["file"],
            "CSV File": csv["file"],
            "Δ Time (s)": round(diff, 1),
            "Signal Similarity": sig,
            "Confidence": conf,
            "Match": ryg(conf)
        })

    return pd.DataFrame(rows), matches

# ==============================
# TIMELINE + SIGNAL OVERLAY
# ==============================
def timeline_with_signals(csv_runs, txt_runs, matches, selected_pairs):

    fig = go.Figure()
    y = 0
    pos = {}

    # Run bars
    for r in csv_runs:
        fig.add_bar(x=[(r["end"] - r["start"]).seconds], y=[y],
                    base=r["start"], orientation="h",
                    marker=dict(color="blue"))
        pos[r["file"]] = y
        y += 1

    for r in txt_runs:
        fig.add_bar(x=[(r["end"] - r["start"]).seconds], y=[y],
                    base=r["start"], orientation="h",
                    marker=dict(color="orange"))
        pos[r["file"]] = y
        y += 1

    # Signal overlays
    for m in matches:
        for c, t in selected_pairs:
            if c in m["csv"]["signals"] and t in m["txt"]["signals"]:

                a = m["csv"]["signals"][c]["values"]
                b = m["txt"]["signals"][t]

                L = min(len(a), len(b))
                x = list(range(L))

                fig.add_trace(go.Scatter(x=x, y=a[:L],
                                         name=f"CSV {c}",
                                         line=dict(color="blue")))

                fig.add_trace(go.Scatter(x=x, y=b[:L],
                                         name=f"TXT {t}",
                                         line=dict(color="orange", dash="dot")))

    return fig

# ==============================
# EXCEL EXPORT
# ==============================
def create_excel(df):
    wb = Workbook()
    ws = wb.active

    colors = {
        "green": PatternFill(start_color="C6EFCE", fill_type="solid"),
        "yellow": PatternFill(start_color="FFEB9C", fill_type="solid"),
        "red": PatternFill(start_color="FFC7CE", fill_type="solid")
    }

    for c, col in enumerate(df.columns, 1):
        ws.cell(1, c, col)

    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            if df.columns[c - 1] == "Confidence":
                if val >= 75:
                    cell.fill = colors["green"]
                elif val >= 40:
                    cell.fill = colors["yellow"]
                else:
                    cell.fill = colors["red"]

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ==============================
# UI
# ==============================
st.title("EOL Matching Dashboard")

with st.expander("ℹ️ Help"):
    st.markdown("""
Upload CSV + TXT → Run analysis  

• Timeline shows run overlap  
• Table shows match confidence  
• Signal plots show behavior comparison  
• Signal mapping allows manual + AI matching  

🟢 Strong | 🟡 Partial | 🔴 Weak  
""")

csv_files = st.file_uploader("CSV files", accept_multiple_files=True)
txt_files = st.file_uploader("TXT files", accept_multiple_files=True)

# ==============================
# RUN
# ==============================
if st.button("Run Analysis"):
    csv_runs = [r for f in csv_files if (r := parse_csv(f))]
    txt_runs = [r for f in txt_files if (r := parse_txt(f))]

    df, matches = run(csv_runs, txt_runs)

    st.session_state.update({
        "df": df, "matches": matches,
        "csv": csv_runs, "txt": txt_runs
    })

# ==============================
# SIGNAL MAPPING
# ==============================
if "csv" in st.session_state and "txt" in st.session_state:

    st.subheader("🔗 Signal Mapping")

    csv_signals = sorted({s for r in st.session_state["csv"] for s in r["signals"]})
    txt_signals = sorted({s for r in st.session_state["txt"] for s in r["signals"]})

    if st.button("🤖 Auto-Suggest Mappings"):
        st.session_state["signal_map"] = suggest_mappings(csv_signals, txt_signals)

    mapping_df = pd.DataFrame({
        "CSV Signal": csv_signals,
        "Mapped TXT Signal": [
            st.session_state.get("signal_map", {}).get(c, "") for c in csv_signals
        ]
    })

    edited = st.data_editor(mapping_df, use_container_width=True)

    if st.button("✅ Apply Mapping"):
        st.session_state["signal_map"] = {
            row["CSV Signal"]: row["Mapped TXT Signal"]
            for _, row in edited.iterrows()
            if row["Mapped TXT Signal"] in txt_signals
        }

    # Save / Load JSON
    json_data = json.dumps(st.session_state.get("signal_map", {}), indent=2)
    st.download_button("💾 Save Mapping", json_data, "mapping.json")

    uploaded = st.file_uploader("Load Mapping", type=["json"])
    if uploaded:
        st.session_state["signal_map"] = json.load(uploaded)

# ==============================
# RESULTS
# ==============================
if "df" in st.session_state:

    df = st.session_state["df"]

    st.subheader("Results")
    st.dataframe(df, use_container_width=True)

    st.download_button("⬇️ Excel Report", create_excel(df), "results.xlsx")

    st.subheader("Timeline + Signal Overlay")

    idx = st.selectbox("Select Match", range(len(df)))
    m = st.session_state["matches"][idx]

    pairs = m["pairs"]
    defaults = best_signals(m["csv"]["signals"], m["txt"]["signals"], pairs)

    selected = st.multiselect(
        "Select Signals",
        [f"{c} ↔ {t}" for c, t in pairs],
        default=[f"{c} ↔ {t}" for c, t in defaults]
    )

    selected_pairs = [tuple(s.split(" ↔ ")) for s in selected]

    fig = timeline_with_signals(
        st.session_state["csv"],
        st.session_state["txt"],
        st.session_state["matches"],
        selected_pairs
    )

    st.plotly_chart(fig)